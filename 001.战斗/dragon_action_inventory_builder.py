import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(r"E:\zhanggz_combat\Proj_Metahuman\Content\M98\Art\Character\Monster\Dragon")
OUT = Path(r"E:\laozhang0101_practice\practice\001.战斗\dragon_action_inventory.csv")
OUT_XLSX = Path(r"E:\laozhang0101_practice\practice\001.战斗\dragon_action_inventory.xlsx")
OUT_MD = Path(r"E:\laozhang0101_practice\practice\001.战斗\Dragon 动作资源识别表.md")
CHUNK_PREFIX = Path(r"E:\laozhang0101_practice\practice\001.战斗\dragon_action_chunk_")


def classify_mesh(name: str) -> tuple[str, str, str, str]:
    if name == "dragon.uasset":
        return ("SkeletalMesh", "非动作资源：龙怪主体模型", "高", "Mesh目录与资源命名")
    if name == "dragon_PhysicsAsset.uasset":
        return ("PhysicsAsset", "非动作资源：龙怪物理碰撞/受击骨骼配置", "高", "PhysicsAsset命名")
    if name == "dragon_Skeleton.uasset":
        return ("Skeleton", "非动作资源：龙怪骨骼资源，供动画驱动", "高", "Skeleton命名")
    if name.startswith("Mat_body_"):
        return ("Material", "非动作资源：身体材质变体", "高", "材质命名")
    if name == "Mat_Eye.uasset":
        return ("Material", "非动作资源：眼球材质", "高", "材质命名")
    if name == "Mat_EyeCover.uasset":
        return ("Material", "非动作资源：眼部覆盖材质", "高", "材质命名")
    if name == "Mat_EyeLenz.uasset":
        return ("Material", "非动作资源：眼睛镜片材质", "高", "材质命名")
    if name == "Mat_PartsBreak.uasset":
        return ("Material", "非动作资源：部件破损材质", "高", "材质命名")
    if name == "Mat_Wet.uasset":
        return ("Material", "非动作资源：湿润效果材质", "高", "材质命名")
    if name.startswith("Mat_Wing_Alpha_"):
        return ("Material", "非动作资源：翅膀透明/遮罩材质", "高", "材质命名")
    return ("OtherMeshAsset", "非动作资源：Mesh目录下其他资源", "中", "目录位置")


def classify_animation(name: str) -> tuple[str, str, str, str]:
    match = re.match(r"^em0002_(\d{2})_(\d{3})(?:_(\d{3}))?(_Loop)?\.uasset$", name)
    if not match:
        return ("Unknown", "未识别命名规则，需人工预览确认", "低", "未知命名")

    group = match.group(1)
    is_loop = name.endswith("_Loop.uasset")
    if is_loop:
        return (
            "AnimationLoop",
            f"动画组{group}的循环动作资源；大概率用于待机/移动/飞行/盘旋/持续咆哮等持续状态，具体动作需在Unreal预览确认",
            "低",
            "文件名含Loop，命名模式em0002_xx_xxx",
        )
    return (
        "AnimationClip",
        f"动画组{group}的一次性动作资源；大概率用于攻击/受击/转向/起落/技能/衔接等单次表现，具体动作需在Unreal预览确认",
        "低",
        "命名模式em0002_xx_xxx",
    )


def main() -> None:
    rows: list[list[str]] = []
    for path in sorted(ROOT.rglob("*")):
        if not path.is_file():
            continue
        rel = str(path.relative_to(ROOT)).replace("/", "\\")
        name = path.name
        if rel.startswith("Mesh\\"):
            resource_type, action, confidence, basis = classify_mesh(name)
        else:
            resource_type, action, confidence, basis = classify_animation(name)
        rows.append([name, rel, resource_type, action, confidence, basis])

    with OUT.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(["文件名称", "相对路径", "资源类型", "动作内容", "动作语义置信度", "识别依据"])
        writer.writerows(rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dragon动作资源"
    headers = ["文件名称", "相对路径", "资源类型", "动作内容", "动作语义置信度", "识别依据"]
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        1: 28,
        2: 52,
        3: 18,
        4: 72,
        5: 16,
        6: 28,
    }
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment
    for idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(idx)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment
    sheet.freeze_panes = "A2"
    workbook.save(OUT_XLSX)

    md_lines = [
        "## 说明",
        "",
        "- 本表基于文件命名规则识别资源类型与动作语义。",
        "- `AnimationLoop` / `AnimationClip` 能确认是动画资源，但具体动作内容仍需在 Unreal 预览中最终确认。",
        "- `Mesh` 目录下的模型、骨骼、物理资产、材质资源已按命名进行直接描述。",
        "",
        "## 资源清单",
        "",
        "| 文件名称 | 相对路径 | 资源类型 | 动作内容 | 动作语义置信度 | 识别依据 |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for row in rows:
        escaped = [cell.replace("|", "\\|").replace("\n", "<br/>") for cell in row]
        md_lines.append(f"| {' | '.join(escaped)} |")
    OUT_MD.write_text("\n".join(md_lines), encoding="utf-8")

    chunk_size = 120
    table_header = [
        "| 文件名称 | 相对路径 | 资源类型 | 动作内容 | 动作语义置信度 | 识别依据 |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for idx in range(0, len(rows), chunk_size):
        chunk_rows = rows[idx : idx + chunk_size]
        chunk_lines = [f"## 资源清单 Part {idx // chunk_size + 1}", "", *table_header]
        for row in chunk_rows:
            escaped = [cell.replace("|", "\\|").replace("\n", "<br/>") for cell in row]
            chunk_lines.append(f"| {' | '.join(escaped)} |")
        chunk_path = Path(f"{CHUNK_PREFIX}{idx // chunk_size + 1}.md")
        chunk_path.write_text("\n".join(chunk_lines), encoding="utf-8")

    print(f"rows={len(rows)}")
    print(f"xlsx={OUT_XLSX}")
    print(f"md={OUT_MD}")
    for row in rows[:8]:
        print(row)


if __name__ == "__main__":
    main()
